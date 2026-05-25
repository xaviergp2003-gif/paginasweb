"""
Diagnóstico de URLs Netlify del Excel: detecta 404 y permite re-publicar.
Uso:
  python3 verificar_netlify.py          # solo revisa
  python3 verificar_netlify.py --fix 1  # re-despliega la fila 1 de cada hoja con HTML en dist/
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from deployer import deploy_dist, get_site_status

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
NETLIFY_RE = re.compile(r"https?://([a-z0-9-]+)\.netlify\.app", re.I)


def _check_url(url: str) -> tuple[str, int, str]:
    try:
        r = requests.get(url, timeout=20, allow_redirects=True)
        return url, r.status_code, r.headers.get("Content-Type", "")[:40]
    except requests.RequestException as e:
        return url, -1, str(e)


def main() -> None:
    import os

    parser = argparse.ArgumentParser()
    parser.add_argument("--fix", type=int, metavar="ROW", help="Re-despliega fila N (2=primera empresa)")
    args = parser.parse_args()

    token = os.getenv("NETLIFY_ACCESS_TOKEN")
    if not token:
        log.error("Crea un archivo .env con NETLIFY_ACCESS_TOKEN=tu_token")
        sys.exit(1)

    if not XLSX.exists():
        log.error("No existe %s", XLSX)
        sys.exit(1)

    wb = load_workbook(XLSX)
    ok = bad = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        log.info("\n=== Hoja: %s ===", sheet)
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 2).value
            if not name:
                continue
            url = str(ws.cell(row, 6).value or "")
            m = NETLIFY_RE.search(url)
            if not m:
                continue
            url = f"https://{m.group(1)}.netlify.app"
            _, code, detail = _check_url(url)

            try:
                st = get_site_status(token, url)
                pub = "sí" if st["published"] else "NO"
            except Exception as e:
                st = {"published": False}
                pub = f"error API: {e}"

            if code == 200 and "text/html" in detail.lower():
                ok += 1
                log.info("  OK  fila %d | %s | %s", row, name, url)
            else:
                bad += 1
                log.warning(
                    "  MAL fila %d | %s | %s → HTTP %s (%s) | publicado: %s",
                    row, name, url, code, detail, pub,
                )

            if args.fix and row == args.fix:
                html = ws.cell(row, 8).value
                dist = Path(str(html)).parent if html else None
                if not dist or not dist.is_absolute():
                    dist = ROOT / dist if dist else None
                if not dist or not (dist / "index.html").is_file():
                    from templater import DIST_DIR
                    for cand in DIST_DIR.glob("*"):
                        if (cand / "index.html").is_file():
                            dist = cand
                            break
                if dist and (dist / "index.html").is_file():
                    log.info("  Re-desplegando %s desde %s…", name, dist)
                    new_url = deploy_dist(dist, netlify_url=url)
                    ws.cell(row, 6, new_url)
                    log.info("  ✓ Actualizado: %s", new_url)
                else:
                    log.error("  No hay dist/ con index.html para fila %d", row)

    if args.fix:
        wb.save(XLSX)

    log.info("\n=== Resumen: %d OK, %d con problemas ===", ok, bad)
    if bad:
        log.info(
            "Para reparar una fila: python3 verificar_netlify.py --fix N\n"
            "O todas las pizzerías: python3 regenerar_pizzerias.py"
        )


if __name__ == "__main__":
    main()
