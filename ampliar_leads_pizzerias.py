"""
Busca más pizzerías sin web, genera landing, sube a Netlify y actualiza Excel.
Mantiene leads existentes con URL Netlify y añade filas nuevas.
"""

from __future__ import annotations

import logging
import re
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from content_ai import generate_copy
from deployer import deploy_dist
from scraper import PlaceLead, search_leads
from templater import (
    DIST_DIR,
    has_whatsapp,
    render_landing,
    whatsapp_href,
    whatsapp_message_info,
)
from webs_storage import save_client_web, webs_path_for_excel

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
HEADERS = [
    "Zona / Búsqueda",
    "Empresa",
    "Teléfono",
    "Dirección",
    "Ciudad",
    "URL Web (Netlify)",
    "Enlace WhatsApp",
    "HTML local",
    "Generado",
]

QUERIES = [
    "Pizzerías en Barcelona",
    "Pizzerías en Eixample Barcelona",
    "Pizzerías en Gràcia Barcelona",
    "Pizzerías en Sants Barcelona",
    "Pizzerías en Poblenou Barcelona",
    "Pizzerías en Badalona",
    "Pizzerías en L'Hospitalet de Llobregat",
    "Pizzerías en Gavà",
    "Pizzerías en Castelldefels",
    "Pizzerías en Cornellà de Llobregat",
    "Pizzerías en Viladecans",
    "Pizzerías en Sitges",
    "Pizzerías en Esplugues de Llobregat",
    "Pizzerías en Sant Boi de Llobregat",
    "Pizzerías en Rubí",
    "Pizzerías en Terrassa",
    "Pizzerías en Mataró",
    "Pizzerías en Sabadell",
    "Pizzerías en Vilanova i la Geltrú",
    "Pizzerías en Premià de Mar",
]

NETLIFY_RE = re.compile(r"https?://[a-z0-9-]+\.netlify\.app", re.I)
MAX_NUEVOS = 15
PAUSA_NETLIFY = 22


def _load_existing() -> tuple[list[list], set[str]]:
    if not XLSX.exists():
        return [], set()
    wb = load_workbook(XLSX)
    ws = wb.active
    rows: list[list] = []
    names: set[str] = set()
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        nombre = str(vals[1] or "").strip()
        if not nombre or nombre in HEADERS or re.fullmatch(r"[\d\s+()-]+", nombre):
            continue
        url = wa = html = ts = ""
        for v in vals:
            s = str(v or "")
            if not url and NETLIFY_RE.search(s):
                url = NETLIFY_RE.search(s).group(0)
            if "wa.me" in s:
                wa = s
            if "dist/" in s or "webs/" in s:
                html = s
            if "2026-" in s and len(s) >= 10:
                ts = s
        if not url:
            continue
        ciudad = vals[4] if vals[4] and not str(vals[4]).startswith("http") else ""
        rows.append([vals[0] or "", nombre, vals[2] or "", vals[3] or "", ciudad, url, wa, html, ts])
        names.add(nombre.lower())
    return rows, names


def _normalize_row(row: list) -> list:
    """Asegura 9 columnas alineadas con HEADERS."""
    while len(row) < 9:
        row.append("")
    zona, nombre, tel, dir_, ciudad, url = row[0], row[1], row[2], row[3], row[4], row[5]
    wa, html, ts = row[6], row[7], row[8]
    if url and NETLIFY_RE.search(str(url)):
        pass
    else:
        for i, v in enumerate(row):
            m = NETLIFY_RE.search(str(v or ""))
            if m:
                url = m.group(0)
                break
    if wa and "wa.me" not in str(wa):
        for v in row:
            if v and "wa.me" in str(v):
                wa = v
                break
    return [zona, nombre, tel, dir_, ciudad, url, wa, html, ts]


def _process(lead: PlaceLead, query: str, idx: int) -> list | None:
    log.info("→ %s", lead.name)
    try:
        copy = generate_copy(lead.name, lead.reviews, city=lead.city, category=lead.category)
    except Exception as e:
        log.error("  IA: %s", e)
        return None
    out = DIST_DIR / f"lead-new-{idx}"
    try:
        render_landing(lead=lead, copy=copy, output_dir=out)
    except Exception as e:
        log.error("  HTML: %s", e)
        return None
    try:
        url = deploy_dist(out, site_name=lead.name)
        log.info("  ✓ %s", url)
    except Exception as e:
        log.error("  Netlify: %s", e)
        return None
    try:
        webs_html = save_client_web(lead.name, out)
    except Exception as e:
        log.error("  webs/: %s", e)
        return None
    wa = whatsapp_href(lead.phone, whatsapp_message_info(lead.name)) if has_whatsapp(lead.phone) else ""
    return [
        query,
        lead.name,
        lead.phone or "—",
        lead.address,
        lead.city,
        url,
        wa,
        webs_path_for_excel(webs_html),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]


def main() -> None:
    clean_rows, seen_names = _load_existing()
    clean_rows = [_normalize_row(r) for r in clean_rows]
    log.info("Excel: %d leads con Netlify existentes", len(clean_rows))

    seen_ids: set[str] = set()
    nuevos: list[list] = []
    deploy_i = 0

    for query in QUERIES:
        if len(nuevos) >= MAX_NUEVOS:
            break
        log.info("--- %s ---", query)
        try:
            found = search_leads(query, max_results=12)
        except Exception as e:
            log.error("Scraper: %s", e)
            continue
        for lead in found:
            if len(nuevos) >= MAX_NUEVOS:
                break
            nk = lead.name.lower().strip()
            if nk in seen_names or lead.place_id in seen_ids:
                continue
            seen_names.add(nk)
            seen_ids.add(lead.place_id)
            deploy_i += 1
            row = _process(lead, query, deploy_i)
            if row:
                nuevos.append(row)
            if deploy_i > 0 and len(nuevos) < MAX_NUEVOS:
                time.sleep(PAUSA_NETLIFY)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pizzerías sin web"
    ws.append(HEADERS)
    for row in clean_rows + nuevos:
        ws.append(_normalize_row(row)[:9])
    wb.save(XLSX)

    log.info("=== FIN: %d total (%d nuevos) → %s ===", len(clean_rows) + len(nuevos), len(nuevos), XLSX)
    for row in nuevos:
        log.info("  + %s → %s", row[1], row[5])


if __name__ == "__main__":
    main()
