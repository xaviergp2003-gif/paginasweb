"""
Procesa leads en orden: Netlify + carpeta webs/ + negrita en Excel.
Ritmo: 3 cada 2 minutos, hasta 90 completados (reanudable).
"""

from __future__ import annotations

import json
import logging
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

from deployer import deploy_dist
from netlify_verify import verify_published
from webs_storage import save_client_web, webs_path_for_excel

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
STATE_FILE = ROOT / ".proceso_lote.json"
DIST_DIR = ROOT / "dist"

BATCH_SIZE = 3
PAUSA_ENTRE_LOTES = 120  # 2 minutos entre lotes de 3
PAUSA_ENTRE_DEPLOYS = 40  # máx ~3 deploys/min (límite Netlify)
META_TOTAL = 90
NETLIFY_RE = re.compile(r"https?://[a-z0-9-]+\.netlify\.app", re.I)
DIST_PAT = re.compile(r"(redeploy-\d+|lead-new-\d+|pelu-sitges-\d+|lead-\d+)", re.I)

SHEETS_ORDEN = ["Pizzerías", "Peluquerías Sitges"]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


def _load_env() -> None:
    for f in (ROOT / ".env", ROOT / "nombres.env"):
        if f.is_file():
            load_dotenv(f)
            return
    load_dotenv()


def _is_header_row(ws, row: int, sheet: str) -> bool:
    v = ws.cell(row, 1).value
    if sheet == "Pizzerías" and row == 1:
        return str(v or "").startswith("Zona")
    return False


def _row_done(ws, row: int) -> bool:
    html = str(ws.cell(row, 8).value or "")
    if "webs/" not in html:
        return False
    font = ws.cell(row, 2).font
    return bool(font and font.bold)


def _resolve_dist(html_cell: str | None, sheet: str, row: int) -> Path | None:
    s = str(html_cell or "")
    m = DIST_PAT.search(s)
    if m:
        d = DIST_DIR / m.group(1)
        if (d / "index.html").is_file():
            return d

    if sheet == "Pizzerías" and row >= 2:
        if row <= 13:
            cand = DIST_DIR / f"redeploy-{row - 1}"
        else:
            cand = DIST_DIR / f"lead-new-{row - 13}"
        if (cand / "index.html").is_file():
            return cand

    if sheet == "Peluquerías Sitges":
        idx = max(1, row if row > 2 else 1)
        cand = DIST_DIR / f"pelu-sitges-{idx}"
        if (cand / "index.html").is_file():
            return cand
        cand = DIST_DIR / f"pelu-sitges-{row}"
        if (cand / "index.html").is_file():
            return cand

    return None


def _netlify_url(ws, row: int) -> str | None:
    for col in (6, 5, 7):
        v = str(ws.cell(row, col).value or "")
        m = NETLIFY_RE.search(v)
        if m:
            return m.group(0)
    return None


def _bold_row(ws, row: int) -> None:
    bold = Font(bold=True)
    for col in range(1, 10):
        ws.cell(row, col).font = bold


def _unbold_row(ws, row: int) -> None:
    normal = Font(bold=False)
    for col in range(1, 10):
        ws.cell(row, col).font = normal


def _collect_queue(wb) -> list[dict]:
    items: list[dict] = []
    seen_names: set[str] = set()
    for sheet in SHEETS_ORDEN:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            if _is_header_row(ws, row, sheet):
                continue
            name = str(ws.cell(row, 2).value or "").strip()
            if not name:
                continue
            key = f"{sheet}:{name.lower()}"
            if key in seen_names:
                continue
            seen_names.add(key)
            if _row_done(ws, row):
                continue
            url = _netlify_url(ws, row)
            if not url:
                continue
            dist = _resolve_dist(ws.cell(row, 8).value, sheet, row)
            if not dist:
                log.warning("Sin HTML local: %s (%s fila %d)", name, sheet, row)
                continue
            items.append({
                "sheet": sheet,
                "row": row,
                "name": name,
                "url": url,
                "dist": str(dist),
                "html_prev": ws.cell(row, 8).value,
            })
    return items


def _load_state() -> set[str]:
    if not STATE_FILE.is_file():
        return set()
    data = json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return set(data.get("done_keys", []))


def _save_state(done_keys: set[str], completed: int) -> None:
    STATE_FILE.write_text(
        json.dumps(
            {
                "done_keys": sorted(done_keys),
                "completed": completed,
                "updated": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def _process_one(wb, item: dict) -> bool:
    ws = wb[item["sheet"]]
    row = item["row"]
    dist = Path(item["dist"])
    name = item["name"]
    url = item["url"]

    log.info("→ [%s] %s", item["sheet"], name)
    try:
        new_url = deploy_dist(dist, netlify_url=url)
        check = verify_published(new_url)
        if not check.get("ok"):
            raise RuntimeError(check.get("message") or "Verificación API fallida")
        webs_html = save_client_web(name, dist, existing_html=item.get("html_prev"))
        ws.cell(row, 6, check.get("ssl_url") or new_url)
        ws.cell(row, 8, webs_path_for_excel(webs_html))
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.cell(row, 9, ts)
        _bold_row(ws, row)
        log.info("  ✓ %s | %s (%s bytes)", new_url, webs_html.parent.name, check.get("size"))
        return True
    except Exception as e:
        log.error("  ✗ %s: %s", name, e)
        _unbold_row(ws, row)
        return False


def _generate_more(needed: int) -> None:
    """Genera pizzerías nuevas hasta cubrir huecos (requiere APIs)."""
    if needed <= 0:
        return
    log.info("Faltan %d: lanzando ampliar_leads (máx %d)…", needed, needed)
    import ampliar_leads_pizzerias as amp

    amp.MAX_NUEVOS = min(needed, 53)
    amp.PAUSA_NETLIFY = 40
    amp.main()


def main() -> None:
    _load_env()
    if not XLSX.is_file():
        log.error("No existe %s", XLSX)
        sys.exit(1)

    wb = load_workbook(XLSX)
    done_keys = _load_state()
    queue = _collect_queue(wb)
    queue = [q for q in queue if f"{q['sheet']}:{q['row']}" not in done_keys]

    already_bold = sum(
        1
        for sheet in SHEETS_ORDEN
        if sheet in wb.sheetnames
        for row in range(1, wb[sheet].max_row + 1)
        if not _is_header_row(wb[sheet], row, sheet)
        and _row_done(wb[sheet], row)
    )
    log.info(
        "Cola: %d pendientes | Ya completados (webs+negrita): %d | Meta: %d",
        len(queue),
        already_bold,
        META_TOTAL,
    )

    completed = already_bold
    batch_num = 0

    while completed < META_TOTAL:
        if not queue and completed < META_TOTAL:
            falta = META_TOTAL - completed
            _generate_more(falta)
            wb = load_workbook(XLSX)
            queue = _collect_queue(wb)
            queue = [q for q in queue if f"{q['sheet']}:{q['row']}" not in done_keys]
            if not queue:
                log.error("No hay más leads para procesar (revisa APIs o Excel).")
                break

        if not queue:
            break

        batch = queue[:BATCH_SIZE]
        queue = queue[BATCH_SIZE:]
        batch_num += 1
        log.info("--- Lote %d (%d leads) ---", batch_num, len(batch))

        for i, item in enumerate(batch):
            if completed >= META_TOTAL:
                break
            key = f"{item['sheet']}:{item['row']}"
            if _process_one(wb, item):
                done_keys.add(key)
                completed += 1
            wb.save(XLSX)
            _save_state(done_keys, completed)
            if i < len(batch) - 1:
                time.sleep(PAUSA_ENTRE_DEPLOYS)

        log.info("Progreso: %d / %d", completed, META_TOTAL)

        if completed >= META_TOTAL or (not queue and completed >= META_TOTAL):
            break
        if queue or completed < META_TOTAL:
            log.info("Esperando %ds antes del siguiente lote…", PAUSA_ENTRE_LOTES)
            time.sleep(PAUSA_ENTRE_LOTES)

    wb.save(XLSX)
    _save_state(done_keys, completed)
    log.info("=== FIN: %d / %d en webs/ + negrita en Excel ===", completed, META_TOTAL)


if __name__ == "__main__":
    main()
