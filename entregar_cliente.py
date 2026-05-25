"""
Entrega al cliente: enlace demo + ZIP para Hostinger.

Uso:
  python3 entregar_cliente.py "Restaurante Pizzeria Isabella"
  python3 entregar_cliente.py --fila 2 --hoja "Pizzerías"
  python3 entregar_cliente.py --excel          # todos con webs/
  python3 entregar_cliente.py --solo-zip       # sin subir demo
"""

from __future__ import annotations

import argparse
import logging
import zipfile
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from github_pages import deploy_github_pages, url_responds_ok
from webs_storage import ROOT, WEBS_DIR, client_slug, webs_path_for_excel

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

XLSX = ROOT / "leads_generados.xlsx"
ENTREGAS_DIR = ROOT / "entregas"
ENLACES_FILE = ROOT / "enlaces_demos.txt"

INSTRUCCIONES = """INSTRUCCIONES — Subir la web a Hostinger
============================================

1. Entra en hPanel de Hostinger → Archivos → Administrador de archivos.
2. Abre la carpeta public_html de tu dominio.
3. Sube el archivo index.html de este ZIP (sustituye el que haya si existe).
4. Abre tu dominio en el navegador y comprueba que se ve la página.

Notas:
- Esta maqueta es una propuesta de diseño. Puedes pedir cambios de textos o fotos.
- El enlace de demo (GitHub Pages) es temporal; en producción usarás tu dominio.
- Si no te interesa el proyecto, no hace falta subir nada: cierra la conversación.

Soporte: responde al WhatsApp desde el que recibiste este enlace.
"""


def _load_env() -> None:
    for f in (ROOT / ".env", ROOT / "nombres.env"):
        if f.is_file():
            load_dotenv(f, override=False)


def _find_webs_folder(name: str, html_cell: str | None) -> Path | None:
    if html_cell and "webs/" in str(html_cell):
        p = ROOT / str(html_cell)
        if p.is_file():
            return p.parent
        if p.is_dir() and (p / "index.html").is_file():
            return p
    slug = client_slug(name)
    direct = WEBS_DIR / slug
    if (direct / "index.html").is_file():
        return direct
    for d in sorted(WEBS_DIR.glob(f"{slug}-*")):
        if (d / "index.html").is_file():
            return d
    for d in WEBS_DIR.iterdir():
        if d.is_dir() and slug[:20] in d.name and (d / "index.html").is_file():
            return d
    return None


def _package_zip(folder: Path, slug: str) -> Path:
    ENTREGAS_DIR.mkdir(parents=True, exist_ok=True)
    pkg_dir = ENTREGAS_DIR / slug
    pkg_dir.mkdir(parents=True, exist_ok=True)
    (pkg_dir / "index.html").write_bytes((folder / "index.html").read_bytes())
    (pkg_dir / "INSTRUCCIONES_HOSTINGER.txt").write_text(INSTRUCCIONES, encoding="utf-8")

    zip_path = ENTREGAS_DIR / f"{slug}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(pkg_dir / "index.html", "index.html")
        zf.write(pkg_dir / "INSTRUCCIONES_HOSTINGER.txt", "INSTRUCCIONES_HOSTINGER.txt")
    return zip_path.resolve()


def _whatsapp_text(empresa: str, demo_url: str | None, zip_path: Path) -> str:
    if demo_url:
        intro = (
            f"Hola, le dejo la maqueta de la web de *{empresa}*:\n"
            f"{demo_url.rstrip('/')}\n\n"
            "Revísala cuando pueda. Si le encaja, le paso el archivo para subirla a su Hostinger "
            "con su dominio. Si prefiere cambios, dígame y los ajustamos."
        )
    else:
        intro = (
            f"Hola, le adjunto la maqueta de la web de *{empresa}* (archivo ZIP).\n\n"
            "Si le encaja, puede subirla a Hostinger siguiendo las instrucciones del ZIP. "
            "Si no le interesa, sin problema."
        )
    return intro + f"\n\nArchivo local (para usted): {zip_path.name}"


def entregar_one(
    empresa: str,
    folder: Path,
    *,
    solo_zip: bool = False,
    update_excel: tuple | None = None,
) -> dict:
    slug = client_slug(empresa)
    zip_path = _package_zip(folder, slug)
    demo_url: str | None = None

    if not solo_zip:
        try:
            demo_url = deploy_github_pages(folder, empresa)
            if not url_responds_ok(demo_url):
                log.warning("La URL demo devolvió error HTTP: %s", demo_url)
        except Exception as e:
            log.error("No se pudo publicar demo: %s", e)
            log.info("ZIP listo igualmente: %s", zip_path)

    if update_excel:
        wb, sheet, row = update_excel
        ws = wb[sheet]
        if demo_url:
            _set_url_cell(ws, row, demo_url)
        ws.cell(row, 9, datetime.now().strftime("%Y-%m-%d %H:%M"))

    return {
        "empresa": empresa,
        "demo_url": demo_url,
        "zip": zip_path,
        "folder": folder,
        "mensaje": _whatsapp_text(empresa, demo_url, zip_path),
    }


def _set_url_cell(ws, row: int, url: str) -> None:
    u = url.rstrip("/")
    cell = ws.cell(row, 6, u)
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=u, display=u)
    cell.font = Font(bold=True, underline="single", color="0563C1")


def _update_excel_by_html(wb, html_rel: str, url: str) -> bool:
    updated = False
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 8).value or "").strip() == html_rel.strip():
                _set_url_cell(ws, row, url)
                ws.cell(row, 9, datetime.now().strftime("%Y-%m-%d %H:%M"))
                updated = True
    return updated


def _append_enlace(folder: Path, empresa: str, url: str) -> None:
    line = f"{url.rstrip('/')}\t{empresa}\t{folder.name}\n"
    with ENLACES_FILE.open("a", encoding="utf-8") as f:
        f.write(line)


def _publish_all_webs(wb, solo_zip: bool) -> int:
    folders = sorted(
        d for d in WEBS_DIR.iterdir() if d.is_dir() and (d / "index.html").is_file()
    )
    if not solo_zip and ENLACES_FILE.exists():
        ENLACES_FILE.unlink()
    done = 0
    for folder in folders:
        empresa = folder.name.replace("-", " ").title()
        html_rel = webs_path_for_excel(folder / "index.html")
        if solo_zip:
            r = entregar_one(empresa, folder, solo_zip=True)
        else:
            try:
                demo_url = deploy_github_pages(folder, empresa)
                zip_path = _package_zip(folder, folder.name)
                r = {
                    "empresa": empresa,
                    "demo_url": demo_url,
                    "zip": zip_path,
                    "folder": folder,
                    "mensaje": _whatsapp_text(empresa, demo_url, zip_path),
                }
                _update_excel_by_html(wb, html_rel, demo_url)
                _append_enlace(folder, empresa, demo_url)
                log.info("✓ %s → %s", folder.name, demo_url)
                done += 1
            except Exception as e:
                log.error("✗ %s: %s", folder.name, e)
                continue
        if solo_zip:
            _print_result(r)
            done += 1
    return done


def _print_result(r: dict) -> None:
    print("\n" + "=" * 60)
    print(f"  {r['empresa']}")
    print("=" * 60)
    if r["demo_url"]:
        print(f"  Demo:  {r['demo_url']}")
    else:
        print("  Demo:  (no publicada — configura GITHUB_TOKEN + GITHUB_PAGES_REPO)")
    print(f"  ZIP:   {r['zip']}")
    print(f"  HTML:  {r['folder'] / 'index.html'}")
    print("\n--- Mensaje para WhatsApp ---\n")
    print(r["mensaje"])
    print("\n" + "=" * 60 + "\n")


def _from_excel_row(wb, sheet: str, row: int, solo_zip: bool) -> dict | None:
    ws = wb[sheet]
    name = ws.cell(row, 2).value
    if not name:
        return None
    folder = _find_webs_folder(str(name), ws.cell(row, 8).value)
    if not folder:
        log.warning("Sin webs/ para fila %d: %s", row, name)
        return None
    return entregar_one(
        str(name),
        folder,
        solo_zip=solo_zip,
        update_excel=(wb, sheet, row),
    )


def main() -> None:
    _load_env()
    ap = argparse.ArgumentParser(description="Enlace demo + ZIP Hostinger para el cliente")
    ap.add_argument("empresa", nargs="?", help="Nombre exacto o parcial del Excel")
    ap.add_argument("--fila", type=int, help="Número de fila en Excel (≥2)")
    ap.add_argument("--hoja", default="Pizzerías", help="Hoja del Excel")
    ap.add_argument("--excel", action="store_true", help="Procesar todas las filas con webs/")
    ap.add_argument(
        "--webs",
        action="store_true",
        help="Un enlace GitHub Pages por cada carpeta en webs/ (82 max)",
    )
    ap.add_argument("--solo-zip", action="store_true", help="Solo ZIP, sin subir demo")
    args = ap.parse_args()

    if args.webs:
        wb = load_workbook(XLSX) if XLSX.exists() else None
        n = _publish_all_webs(wb, args.solo_zip)
        if wb:
            wb.save(XLSX)
        log.info("Publicados: %d | Lista: %s", n, ENLACES_FILE)
        return

    if not XLSX.exists():
        log.error("No existe %s", XLSX)
        return

    wb = load_workbook(XLSX)

    if args.excel:
        done = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in range(2, ws.max_row + 1):
                if not ws.cell(row, 2).value:
                    continue
                html = str(ws.cell(row, 8).value or "")
                if "webs/" not in html:
                    continue
                r = _from_excel_row(wb, sheet, row, args.solo_zip)
                if r:
                    _print_result(r)
                    done += 1
        wb.save(XLSX)
        log.info("Entregas completadas: %d (Excel guardado)", done)
        return

    if args.fila:
        r = _from_excel_row(wb, args.hoja, args.fila, args.solo_zip)
        if r:
            wb.save(XLSX)
            _print_result(r)
        return

    if not args.empresa:
        ap.print_help()
        return

    needle = args.empresa.lower()
    found: list[tuple[str, int, Path]] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            name = str(ws.cell(row, 2).value or "")
            if needle in name.lower():
                folder = _find_webs_folder(name, ws.cell(row, 8).value)
                if folder:
                    found.append((sheet, row, folder))
                    break

    if not found:
        log.error('No encontrado en Excel/webs: "%s"', args.empresa)
        return

    sheet, row, folder = found[0]
    name = str(wb[sheet].cell(row, 2).value)
    r = entregar_one(
        name,
        folder,
        solo_zip=args.solo_zip,
        update_excel=(wb, sheet, row),
    )
    wb.save(XLSX)
    _print_result(r)


if __name__ == "__main__":
    main()
