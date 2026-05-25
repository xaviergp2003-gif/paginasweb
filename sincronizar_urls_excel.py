"""
Escribe las URLs de GitHub Pages en leads_generados.xlsx (columna F).

Lee enlaces_demos.txt y/o genera la URL desde el nombre de carpeta en webs/.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from webs_storage import ROOT, client_slug

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

XLSX = ROOT / "leads_generados.xlsx"
ENLACES_FILE = ROOT / "enlaces_demos.txt"
SHEET_TODOS = "Enlaces demos"


def _load_env() -> None:
    for f in (ROOT / ".env", ROOT / "nombres.env"):
        if f.is_file():
            load_dotenv(f, override=False)


def _owner() -> str:
    return os.getenv("GITHUB_OWNER", "xaviergp2003-gif").strip()


def _url_for_folder(folder: str) -> str:
    return f"https://{_owner().lower()}.github.io/{folder}"


def _load_enlaces_map() -> dict[str, str]:
    m: dict[str, str] = {}
    if ENLACES_FILE.is_file():
        for line in ENLACES_FILE.read_text(encoding="utf-8").splitlines():
            parts = line.strip().split("\t")
            if len(parts) >= 3:
                m[parts[2].strip()] = parts[0].strip().rstrip("/")
    for d in (ROOT / "webs").iterdir():
        if d.is_dir() and (d / "index.html").is_file():
            m.setdefault(d.name, _url_for_folder(d.name))
    return m


def _folder_from_html(html: str) -> str | None:
    s = str(html or "").strip().replace("\\", "/")
    if not s.startswith("webs/"):
        return None
    parts = s.split("/")
    if len(parts) >= 2:
        return parts[1]
    return None


def _set_url_cell(ws, row: int, url: str) -> None:
    u = url.rstrip("/")
    cell = ws.cell(row, 6, u)
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=u, display=u)
    cell.font = Font(bold=True, underline="single", color="0563C1")


def _sync_leads_sheets(wb, enlaces: dict[str, str]) -> int:
    updated = 0
    for sheet in wb.sheetnames:
        if sheet == SHEET_TODOS:
            continue
        ws = wb[sheet]
        if ws.max_row >= 1:
            ws.cell(1, 6, "URL Web (GitHub)")
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 2).value
            if not name:
                continue
            html = str(ws.cell(row, 8).value or "")
            folder = _folder_from_html(html)
            url = None
            if folder and folder in enlaces:
                url = enlaces[folder]
            if not url:
                slug = client_slug(str(name))
                for key, u in enlaces.items():
                    if key == slug or key.startswith(f"{slug}-"):
                        url = u
                        folder = key
                        break
            if url:
                _set_url_cell(ws, row, url)
                if folder and html and folder not in html:
                    ws.cell(row, 8, f"webs/{folder}/index.html")
                ws.cell(row, 9, datetime.now().strftime("%Y-%m-%d %H:%M"))
                updated += 1
            else:
                log.warning("Sin enlace: %s (html=%s)", name, html or "—")
    return updated


def _write_todos_sheet(wb, enlaces: dict[str, str]) -> None:
    if SHEET_TODOS in wb.sheetnames:
        del wb[SHEET_TODOS]
    ws = wb.create_sheet(SHEET_TODOS)
    ws.append(["Carpeta webs", "URL GitHub Pages", "Empresa (texto)"])
    for folder in sorted(enlaces.keys()):
        url = enlaces[folder]
        empresa = folder.replace("-", " ").title()
        ws.append([folder, url, empresa])
        cell = ws.cell(ws.max_row, 2)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url, display=url)
        cell.font = Font(underline="single", color="0563C1")


def main() -> None:
    _load_env()
    if not XLSX.exists():
        log.error("No existe %s", XLSX)
        return
    enlaces = _load_enlaces_map()
    log.info("Enlaces disponibles: %d", len(enlaces))
    wb = load_workbook(XLSX)
    n = _sync_leads_sheets(wb, enlaces)
    _write_todos_sheet(wb, enlaces)
    wb.save(XLSX)
    log.info("Excel actualizado: %d filas en Pizzerías/Peluquerías, hoja '%s' con %d enlaces", n, SHEET_TODOS, len(enlaces))
    log.info("Cierra y vuelve a abrir el Excel si lo tenías abierto.")


if __name__ == "__main__":
    main()
