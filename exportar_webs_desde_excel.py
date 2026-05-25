"""
Copia a webs/{cliente}/ el HTML de filas que ya tienen URL Netlify en el Excel.
Útil para leads antiguos que solo tenían dist/.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from webs_storage import save_client_web, webs_path_for_excel

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
NETLIFY_RE = re.compile(r"https?://[a-z0-9-]+\.netlify\.app", re.I)


def _html_source(row_vals: list, name: str, row: int) -> Path | None:
    for v in row_vals:
        s = str(v or "")
        if "index.html" in s:
            p = Path(s)
            if not p.is_absolute():
                p = ROOT / p
            if p.is_file():
                return p
            if p.parent.is_dir() and (p.parent / "index.html").is_file():
                return p.parent / "index.html"
    from templater import DIST_DIR

    for pattern in (f"lead-{row - 1}", f"lead-new-{row - 1}", f"pelu-sitges-{row - 1}"):
        d = DIST_DIR / pattern
        if (d / "index.html").is_file():
            return d / "index.html"
    slug = re.sub(r"[^a-z0-9-]", "", name.lower().replace(" ", "-"))[:40]
    for d in sorted((ROOT / "dist").glob("*")):
        if d.is_dir() and (d / "index.html").is_file() and slug[:15] in d.name.lower():
            return d / "index.html"
    return None


def main() -> None:
    if not XLSX.exists():
        log.error("No existe %s", XLSX)
        return
    wb = load_workbook(XLSX)
    done = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        log.info("Hoja: %s", sheet)
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 2).value
            if not name:
                continue
            vals = [ws.cell(row, c).value for c in range(1, 10)]
            url = str(vals[5] or "")
            if not NETLIFY_RE.search(url):
                for v in vals:
                    m = NETLIFY_RE.search(str(v or ""))
                    if m:
                        url = m.group(0)
                        break
            if not NETLIFY_RE.search(url):
                continue
            html_path = ws.cell(row, 8).value
            if html_path and "webs/" in str(html_path):
                p = ROOT / str(html_path) if not Path(str(html_path)).is_absolute() else Path(str(html_path))
                if p.is_file():
                    log.info("  ya en webs: %s", name)
                    continue
            src = _html_source(vals, str(name), row)
            if not src:
                log.warning("  sin HTML: %s", name)
                continue
            try:
                dest = save_client_web(str(name), src, existing_html=html_path)
                ws.cell(row, 8, webs_path_for_excel(dest))
                done += 1
                log.info("  ✓ %s → %s", name, dest.parent.name)
            except Exception as e:
                log.error("  %s: %s", name, e)
    wb.save(XLSX)
    log.info("=== Exportados: %d ===", done)


if __name__ == "__main__":
    main()
