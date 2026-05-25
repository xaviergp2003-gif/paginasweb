"""
Diagnóstico: claves API + estado real de URLs Netlify (vía API, no navegador).
Uso: python3 diagnostico_claves.py
"""

from __future__ import annotations

import re
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from netlify_verify import NETLIFY_RE, check_all_keys, verify_published

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"


def _load_env() -> None:
    for f in (ROOT / ".env", ROOT / "nombres.env"):
        if f.is_file():
            load_dotenv(f)
            return
    load_dotenv()


def main() -> None:
    _load_env()
    print("=== CLAVES API ===")
    for k, v in check_all_keys().items():
        print(f"  {k}: {v}")

    if not XLSX.is_file():
        print("\nNo hay Excel.")
        return

    print("\n=== FILAS EN NEGRITA (verificación API) ===")
    wb = load_workbook(XLSX)
    ok = fail = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            font = ws.cell(row, 2).font
            if not (font and font.bold):
                continue
            name = ws.cell(row, 2).value
            url = str(ws.cell(row, 6).value or "")
            m = NETLIFY_RE.search(url)
            if not m:
                print(f"  ✗ {sheet} r{row} {name}: sin URL Netlify")
                fail += 1
                continue
            url = f"https://{m.group(1)}.netlify.app"
            r = verify_published(url)
            if r.get("ok"):
                ok += 1
                print(f"  ✓ {name} → {r['ssl_url']} ({r['size']} bytes)")
            else:
                fail += 1
                print(f"  ✗ {name} → {r.get('message')}")

    print(f"\nResumen negrita: {ok} OK, {fail} con problema")
    if ok and fail == 0:
        print(
            "\nSi en el navegador ves error pero aquí sale OK: suele ser "
            "rate limit temporal de Netlify (429). Espera 15–30 min y recarga."
        )


if __name__ == "__main__":
    main()
