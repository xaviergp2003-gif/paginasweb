"""
Regenera HTML con la plantilla actual y re-despliega las pizzerías de la hoja Excel.
"""

from __future__ import annotations

import json
import logging
import time
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from content_ai import generate_copy
from deployer import deploy_dist
from scraper import PlaceLead, search_leads
from templater import DIST_DIR, render_landing

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
SHEET = "Pizzerías sin web" # Sheet created by ampliar_leads.py
PAUSA = 22


def _find_lead(name: str, city: str = "") -> PlaceLead | None:
    for q in (f"{name} {city}".strip(), name):
        hits = search_leads(q, max_results=5)
        for h in hits:
            if name.lower()[:12] in h.name.lower() or h.name.lower()[:12] in name.lower():
                return h
        if hits:
            return hits[0]
    return None


def _load_copy(path: Path) -> dict | None:
    f = path / "copy.json"
    if f.is_file():
        return json.loads(f.read_text(encoding="utf-8"))
    return None


def _save_copy(path: Path, copy: dict) -> None:
    (path / "copy.json").write_text(json.dumps(copy, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    wb = load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        # Check if the active sheet or "Pizzerías" exists
        if "Pizzerías" in wb.sheetnames:
            ws = wb["Pizzerías"]
        else:
            ws = wb.active
    else:
        ws = wb[SHEET]
        
    updated = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        city = ws.cell(row, 5).value or ""
        url = ws.cell(row, 6).value or ""
        html_path = ws.cell(row, 8).value or ""
        if not name:
            continue
            
        dist = Path(str(html_path)).parent if html_path else DIST_DIR / f"lead-{row - 1}"
        if not dist.is_absolute():
            dist = ROOT / dist
        log.info("→ %s", name)

        lead = _find_lead(str(name), str(city))
        if not lead:
            log.warning("  No encontrado en Places, omitido")
            continue

        copy = _load_copy(dist)
        if not copy:
            try:
                copy = generate_copy(lead.name, lead.reviews, city=lead.city or str(city), category=lead.category)
                _save_copy(dist, copy)
            except Exception as e:
                log.error("  IA: %s", e)
                continue

        try:
            render_landing(lead=lead, copy=copy, output_dir=dist)
        except Exception as e:
            log.error("  HTML: %s", e)
            continue

        if url and "netlify.app" in str(url):
            try:
                new_url = deploy_dist(dist, netlify_url=str(url).strip())
                ws.cell(row, 6, new_url)
                ws.cell(row, 8, str(dist / "index.html"))
                updated += 1
                log.info("  ✓ %s", new_url)
            except Exception as e:
                log.error("  Netlify: %s", e)
        else:
            log.info("  HTML local OK (sin URL Netlify)")
            updated += 1

        time.sleep(PAUSA)

    wb.save(XLSX)
    log.info("=== Regeneradas: %d ===", updated)


if __name__ == "__main__":
    main()
