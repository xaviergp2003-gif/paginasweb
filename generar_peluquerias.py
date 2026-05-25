"""
Peluquerías sin web en Sitges → landing salón → Netlify → hoja Excel nueva.
No modifica la hoja de Pizzerías.
"""

from __future__ import annotations

import logging
import re
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from content_ai import generate_copy_peluqueria
from deployer import deploy_dist
from scraper import PlaceLead, search_leads
from templater import (
    DIST_DIR,
    has_whatsapp,
    render_peluqueria_landing,
    whatsapp_href,
    whatsapp_message_info,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "leads_generados.xlsx"
SHEET_PELU = "Peluquerías Sitges"
HEADERS = [
    "Zona / Búsqueda",
    "Empresa",
    "Teléfono",
    "Dirección",
    "Ciudad",
    "URL Web (Netlify)",
    "Enlace WhatsApp",
    "HTML local",
    "Generado",
]

QUERIES = [
    "Peluquerías en Sitges",
    "Peluquería en Sitges",
    "Salón de belleza Sitges",
    "Peluquería unisex Sitges",
    "Barbería Sitges",
]

MAX_LEADS = 10
PAUSA = 22

# Excluir negocios que no son peluquería
SKIP_NAME = re.compile(
    r"restaurante|pizzer|hotel|farmacia|supermerc|gimnasio|dentista|veterinar|"
    r"abogad|inmobiliar|taller|mecánic|ferreter",
    re.I,
)


def _is_hair_salon(lead: PlaceLead) -> bool:
    text = f"{lead.name} {lead.category}".lower()
    if SKIP_NAME.search(text):
        return False
    return any(
        w in text
        for w in (
            "peluquer", "peluquería", "hair", "salon", "salón", "barber",
            "barbershop", "estilista", "beauty", "belleza", "coiff",
        )
    )


def _append_sheet(rows: list[list]) -> None:
    wb = load_workbook(XLSX)
    if SHEET_PELU in wb.sheetnames:
        ws = wb[SHEET_PELU]
        start = ws.max_row + 1
    else:
        ws = wb.create_sheet(SHEET_PELU)
        for s in wb.sheetnames:
            if s != SHEET_PELU:
                wb[s].sheet_view.tabSelected = False
        ws.sheet_view.tabSelected = True
        ws.append(HEADERS)
        start = 2
    for i, row in enumerate(rows):
        for c, val in enumerate(row[:9], 1):
            ws.cell(start + i - 1, c, val)
    if "Pizzerías" not in wb.sheetnames and wb.sheetnames:
        wb[wb.sheetnames[0]].title = "Pizzerías"
    wb.save(XLSX)


def main() -> None:
    log.info("=== Peluquerías Sitges (sin web) ===")
    DIST_DIR.mkdir(parents=True, exist_ok=True)
    seen: set[str] = set()
    results: list[list] = []
    deploy_i = 0

    for query in QUERIES:
        if len(results) >= MAX_LEADS:
            break
        log.info("--- %s ---", query)
        try:
            found = search_leads(query, max_results=15)
        except Exception as e:
            log.error("Scraper: %s", e)
            continue

        for lead in found:
            if len(results) >= MAX_LEADS:
                break
            key = lead.name.lower().strip()
            if key in seen or lead.place_id in seen:
                continue
            if not _is_hair_salon(lead):
                log.info("  omitido (no peluquería): %s", lead.name)
                continue
            seen.add(key)
            seen.add(lead.place_id)

            log.info("→ %s", lead.name)
            try:
                copy = generate_copy_peluqueria(
                    lead.name, lead.reviews, city=lead.city or "Sitges"
                )
            except Exception as e:
                log.error("  IA: %s", e)
                continue

            deploy_i += 1
            out = DIST_DIR / f"pelu-sitges-{deploy_i}"
            try:
                render_peluqueria_landing(lead=lead, copy=copy, output_dir=out)
            except Exception as e:
                log.error("  HTML: %s", e)
                continue

            try:
                url = deploy_dist(out, site_name=f"peluqueria-{lead.name}")
            except Exception as e:
                log.error("  Netlify: %s", e)
                url = "Pendiente Netlify"

            wa = (
                whatsapp_href(lead.phone, whatsapp_message_info(lead.name))
                if has_whatsapp(lead.phone)
                else ""
            )
            results.append([
                query,
                lead.name,
                lead.phone or "—",
                lead.address,
                lead.city or "Sitges",
                url,
                wa,
                str(out / "index.html"),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ])
            log.info("  ✓ %s", url)
            if len(results) < MAX_LEADS:
                time.sleep(PAUSA)

    if results:
        _append_sheet(results)

    log.info("=== FIN: %d peluquerías → hoja '%s' ===", len(results), SHEET_PELU)
    for r in results:
        log.info("  • %s → %s", r[1], r[5])


if __name__ == "__main__":
    main()
